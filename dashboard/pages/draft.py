"""
Page 3 — Newsletter Draft.

Shows every categorised article (action ∈ {accept_top1, accept_top2, manual}),
grouped by section. Curator can:
  - edit each summary in place + Save
  - click Generate summary to re-draft via Claude
  - click × in the top-right of a card to remove from the draft (sets reject)
  - download the lot as Excel (TITLE / SOURCE / URL / SUMMARY / SECTION / DATE)
  - leave free-text feedback at the bottom (→ curator_feedback.suggestions)
"""

from datetime import datetime
from io import BytesIO

import streamlit as st
import pandas as pd

from dashboard.config import CATEGORY_LABELS, CATEGORY_ORDER, SOURCE_LABELS
from dashboard.data import (
    archive_and_reset_week, clean_text, fetch_article_text, get_accepted_articles,
    is_authenticated, load_decisions, record_decision, record_feedback,
    record_summary,
)
from src.inference.summarise import summarise_article
from src.inference.summarise import summarise_article


def _category_of(art: dict) -> str | None:
    """Effective section assignment — what Page 2 set on the curator_decisions row."""
    return art.get("curator_label") or art.get("top1")


def _build_excel(grouped: dict, today: datetime) -> bytes:
    """Single-sheet Excel matching the curator MS Form export structure, so the
    download can be merged directly with the curator-suggestions spreadsheet.

    Columns match the MS Form export (ERPNewsletterSubmissions.xlsx) EXACTLY —
    same names and order — so the download is the same format and can be pasted
    straight into the form's spreadsheet:
      Id, Start time, Completion time, Email, Name, Organisation, Title,
      Include, Link (website address / URL), Short description,
      Which section of the newsletter is this for?, Any other comments?,
      Submitter, Question?

    The Form auto-fields (Completion time, Email, Name, Question?) are blank for
    scraped articles. "Include" defaults to "Yes" (the curator accepted all of
    these on Page 2). "Submitter" comes from the per-article dropdown on the
    Draft page.
    """
    rows = []
    row_id = 1
    for cat_key in CATEGORY_ORDER:
        for art in grouped.get(cat_key, []):
            url = art.get("url") or ""
            session_key = f"desc_{url}"
            summary = (
                st.session_state.get(session_key)
                or clean_text(art.get("summary"))
                or "Summary unavailable"
            )
            src = clean_text(art.get("source"))
            article_date = art.get("article_date") or ""
            # Render article_date as DD/MM/YYYY to match MS Forms convention
            try:
                d = pd.to_datetime(article_date, errors="coerce", dayfirst=True)
                date_str = d.strftime("%d/%m/%Y") if not pd.isna(d) else (article_date or "")
            except Exception:
                date_str = article_date or ""
            # EXACT column set + order of the live MS Form export
            # (ERPNewsletterSubmissions.xlsx) so the download is the same format
            # and can be pasted straight into the form's spreadsheet. The Form
            # auto-fields (Completion time / Email / Name / Question?) don't
            # apply to scraped articles, so they're left blank.
            rows.append({
                "Id": row_id,
                "Start time": date_str,
                "Completion time": "",
                "Email": "",
                "Name": "",
                "Organisation": SOURCE_LABELS.get(src, src),
                "Title": clean_text(art.get("title")),
                "Include": "Yes",
                "Link (website address / URL)": url,
                "Short description": summary,
                "Which section of the newsletter is this for?": CATEGORY_LABELS.get(cat_key, cat_key),
                "Any other comments?": st.session_state.get(f"notes_{url}", "") or "",
                "Submitter": st.session_state.get(f"submitter_{url}", "") or "",
                "Question?": "",
            })
            row_id += 1
    columns = [
        "Id", "Start time", "Completion time", "Email", "Name",
        "Organisation", "Title", "Include", "Link (website address / URL)",
        "Short description", "Which section of the newsletter is this for?",
        "Any other comments?", "Submitter", "Question?",
    ]
    df = pd.DataFrame(rows, columns=columns)
    buf = BytesIO()

    # Formatted Excel — bold header row, ESRC navy fill on headers,
    # auto-width per column (capped), wrapped text on long-content columns,
    # frozen header row, banded rows.
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Newsletter draft")
        ws = writer.sheets["Newsletter draft"]

        # Header styling
        header_fill = PatternFill("solid", fgColor="0F1E3D")  # ESRC navy
        header_font = Font(bold=True, color="FFFFFF", size=11)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin = Side(border_style="thin", color="888888")
        cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = cell_border

        # Per-column widths (chars) + wrap settings
        wrap_cols = {"Title", "Short description", "Any other comments?"}
        width_map = {
            "Id": 6, "Start time": 13, "Organisation": 22, "Title": 50,
            "Include": 9, "Link (website address / URL)": 60,
            "Short description": 60,
            "Which section of the newsletter is this for?": 28,
            "Any other comments?": 30, "Submitter": 14,
        }
        for col_idx, col_name in enumerate(columns, start=1):
            letter = get_column_letter(col_idx)
            ws.column_dimensions[letter].width = width_map.get(col_name, 18)

        # Body cell styling — wrap text on long-text columns, vertical-top
        # alignment everywhere so wrapped content stays readable.
        wrap_align = Alignment(wrap_text=True, vertical="top")
        plain_align = Alignment(vertical="top")
        for row_idx in range(2, ws.max_row + 1):
            # Banded zebra fill on even rows (subtle)
            band_fill = PatternFill("solid", fgColor="F4F6FA") if row_idx % 2 == 0 else None
            for col_idx, col_name in enumerate(columns, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = wrap_align if col_name in wrap_cols else plain_align
                if band_fill:
                    cell.fill = band_fill

        # Header row taller; freeze it
        ws.row_dimensions[1].height = 32
        ws.freeze_panes = "A2"

    buf.seek(0)
    return buf.getvalue()


def _render_new_week_button():
    """'Start a new week' archive + reset control. Rendered whether or not the
    draft has articles, so the curator can always find it (including right after
    a reset, when the draft is empty). Login-gated + checkbox-confirmed.

    Non-destructive: archives this week's kept/categorised decisions to the
    database, then sets a week boundary so they drop out of Categorise + Draft.
    Kept/rejected articles keep their status (won't reappear in Review);
    pending articles are untouched."""
    if not is_authenticated():
        return
    st.markdown("---")
    st.markdown("### 🗓️ Start a new week")
    st.caption(
        "Archives this week's articles (a copy is saved to the database) and "
        "clears them from Categorise + Draft so next week starts fresh. "
        "Kept/rejected articles will NOT reappear in Review; un-actioned "
        "(pending) articles are untouched. **Download the Excel above first** "
        "as your own copy."
    )
    confirm = st.checkbox(
        "I've downloaded the Excel and want to start a new week",
        key="_reset_confirm",
    )
    if st.button(
        "Archive & start new week", disabled=not confirm,
        use_container_width=True, key="_reset_week_btn",
    ):
        week_label = f"week up to {datetime.now().strftime('%a %d %b %Y')}"
        with st.spinner("Archiving and resetting…"):
            result = archive_and_reset_week(week_label)
        st.session_state["_reset_confirm"] = False
        st.success(
            f"Archived {result['archived']} article(s) as '{week_label}'. "
            "New week started — Categorise and Draft are now clear."
        )
        st.rerun()


def render(df):
    st.title("Step 3: Draft Newsletter")

    accepted = get_accepted_articles(df)
    if not accepted:
        st.info(
            "No categorised articles yet. Use **Triage** → **Select Categories** "
            "first, then come back."
        )
        _render_new_week_button()
        return

    # Group by curator-assigned category. Articles without a valid category
    # are silently skipped — Page 2 always sets one, so this shouldn't happen.
    grouped: dict[str, list[dict]] = {k: [] for k in CATEGORY_ORDER}
    for art in accepted:
        cat = _category_of(art)
        if cat in grouped:
            grouped[cat].append(art)

    today = datetime.now()
    newsletter_date = today.strftime("%d %B %Y")

    # Newsletter header
    st.markdown(
        f"""
        <div style="background:#0f1e3d;color:white;padding:20px;border-radius:8px;
                    text-align:center;margin-bottom:20px;">
            <div style="font-size:18px;font-weight:700;">ESRC Education Research Programme</div>
            <div style="font-size:14px;color:#8fa8c8;margin-top:4px;">ERP Newsletter</div>
            <div style="font-size:14px;color:#8fa8c8;">{newsletter_date}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    auth = is_authenticated()
    decisions = load_decisions()

    # ── Articles by section ─────────────────────────────────────────────────
    for cat_key in CATEGORY_ORDER:
        articles = grouped.get(cat_key, [])
        if not articles:
            continue
        cat_label = CATEGORY_LABELS[cat_key]
        st.markdown(
            f"<div style='background:#1d3461;padding:10px 16px;border-radius:4px;"
            f"margin:18px 0 8px 0;'>"
            f"<span style='color:#c8d8ec;font-size:15px;font-weight:600;'>{cat_label}</span></div>",
            unsafe_allow_html=True,
        )

        for art in articles:
            art_url = art.get("url") or ""
            title = art.get("title") or "No title"
            source_name = SOURCE_LABELS.get(art.get("source", ""), art.get("source") or "")
            article_date = art.get("article_date") or ""
            session_key = f"desc_{art_url}"

            # Seed text_area initial value from DB on first render of this URL.
            # Fallback order: curator edit > pre-generated articles.summary > empty.
            # articles.summary is exposed via v_dashboard after migration 012.
            if session_key not in st.session_state:
                st.session_state[session_key] = (
                    (decisions.get(art_url) or {}).get("summary")
                    or art.get("summary")
                    or ""
                )

            with st.container(border=True):
                # Title / Source / URL in one tight block (no gaps between).
                # X-remove button sits in the top-right corner.
                col_t, col_x = st.columns([10, 1])
                with col_t:
                    st.markdown(
                        f"<p style='margin:0;'><b>TITLE:</b> {title}</p>"
                        f"<p style='margin:0;'><b>SOURCE:</b> {source_name}</p>"
                        + (f"<p style='margin:0 0 6px 0;overflow-wrap:anywhere;'>"
                           f"<b>URL:</b> <a href='{art_url}' target='_blank'>{art_url}</a></p>"
                           if art_url else ""),
                        unsafe_allow_html=True,
                    )
                with col_x:
                    if st.button(
                        "×",
                        key=f"xrm_{art_url}",
                        help="Remove from newsletter (sets to Rejected)",
                        disabled=not auth,
                    ):
                        record_decision(art_url, "reject", "")
                        st.rerun()

                # The newsletter copy = the AI-generated 1-2 sentence summary,
                # editable here. Generate (re)builds it from the article body.
                st.markdown("**SUMMARY:**")
                col_summary, col_save, col_gen = st.columns([6, 1, 1])
                with col_summary:
                    edited_desc = st.text_area(
                        "summary",
                        key=session_key,
                        height=120,
                        label_visibility="collapsed",
                        disabled=not auth,
                    )
                with col_save:
                    st.markdown(
                        "<div style='height:28px'></div>", unsafe_allow_html=True
                    )  # vertical align nudge
                    if st.button(
                        "💾 Save", key=f"sv_{art_url}",
                        type="primary", disabled=not auth,
                    ):
                        record_summary(art_url, edited_desc)
                        st.toast("Saved.")
                with col_gen:
                    st.markdown(
                        "<div style='height:28px'></div>", unsafe_allow_html=True
                    )
                    if st.button(
                        "✎ Generate", key=f"gen_{art_url}", disabled=not auth,
                        help="AI writes a 1-2 sentence summary from the article.",
                    ):
                        with st.spinner("Summarising via Claude…"):
                            body = fetch_article_text(art_url)
                            new_summary = summarise_article(
                                title=title, text=body,
                                category=art.get("curator_label") or art.get("top1"),
                            )
                        record_summary(art_url, new_summary)
                        # Drop the cached text_area value so it re-seeds from the
                        # freshly-written summary on rerun.
                        if session_key in st.session_state:
                            del st.session_state[session_key]
                        st.rerun()

                # Per-article fillable fields — BELOW the summary
                col_section, col_notes, col_submitter = st.columns(3)
                with col_section:
                    # Re-assign the newsletter section in place. Picking a
                    # different category commits immediately via on_change —
                    # records action=manual and the card regroups under the
                    # new section on rerun (same mechanic as Page 2).
                    def _on_change_section(_url=art_url, _current=cat_key):
                        choice = st.session_state.get(f"sect_{_url}")
                        if choice and choice != _current:
                            record_decision(_url, "manual", choice)

                    st.selectbox(
                        "Section",
                        options=list(CATEGORY_ORDER),
                        index=list(CATEGORY_ORDER).index(cat_key),
                        format_func=lambda x: CATEGORY_LABELS.get(x, x),
                        key=f"sect_{art_url}",
                        on_change=_on_change_section,
                        disabled=not auth,
                    )
                with col_notes:
                    st.text_input(
                        "Any other comments?",
                        key=f"notes_{art_url}",
                        placeholder="Optional curator note for this row",
                        disabled=not auth,
                    )
                with col_submitter:
                    st.selectbox(
                        "Submitter",
                        options=["", "GM", "RF", "NC"],
                        key=f"submitter_{art_url}",
                        disabled=not auth,
                    )

    # ── Download Excel ──────────────────────────────────────────────────────
    st.markdown("---")
    excel_bytes = _build_excel(grouped, today)
    st.download_button(
        "📥 Download newsletter as Excel",
        excel_bytes,
        file_name=f"newsletter_{today.strftime('%Y_%m_%d')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        type="primary",
    )
    st.caption(
        "Tip: hit **Save** on any edited summary before downloading — "
        "unsaved edits will still appear in the Excel, but won't persist if you reload."
    )

    # ── Start a new week (archive + reset) ────────────────────────────────────
    _render_new_week_button()

    # Feedback box has been moved to app.py (renders at the bottom of every page).
